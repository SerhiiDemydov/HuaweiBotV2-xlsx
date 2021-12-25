import telebot
from English import english
from Russian import russian
from openpyxl import load_workbook



API_TOKEN = "919816906:AAELsHd-wuK8AZYQwifA1JUz6_ZJTKA60Qw"
keyboardLanguage = telebot.types.ReplyKeyboardMarkup(True,)
keyboardLanguage.row('English/Английский')
keyboardLanguage.row('Russian/Русский')



bot = telebot.TeleBot(API_TOKEN)


name = []
type = []
id = []
Language = []

def read_user_info(name_file,sheet_name):
    global name, type, id, Language
    wb = load_workbook(filename=name_file)
    sheet = wb[sheet_name]
    j = 1
    while True:
        j += 1
        if sheet[f"A{j}"].value:
            id.append(sheet[f"A{j}"].value)
            name.append(sheet[f"B{j}"].value)
            Language.append(sheet[f"C{j}"].value)
            type.append(sheet[f"D{j}"].value)
        else:
            break

def write_user_info(name_file,sheet_name):
    global name, type, id, Language
    wb = load_workbook(filename=name_file)
    sheet = wb[sheet_name]
    for j in range(len(id)):
        sheet[f"A{j+2}"].value = id[j]
        sheet[f"B{j+2}"].value = name[j]
        sheet[f"C{j+2}"].value = Language[j]
        sheet[f"D{j+2}"].value = type[j]
    wb.save(name_file)


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Выбирете язык / Select language',reply_markup=keyboardLanguage)


@bot.message_handler(content_types=['text'])
def send_text1(message):
    global type, id, name, Language
    find = False

    for i in range(len(id)):
        if message.from_user.id == id[i]:
            find = True
            print("User found")
            break


    if find == False:
        print("User didn't find")
        i = 0
        id.append(message.from_user.id)
        name.append(str(message.from_user.first_name)) #+ " " + message.from_user.last_name))
        type.append(0)
        Language.append(0)



    if message.text.lower() == 'english/английский' or message.text.lower() == 'английский':
        Language[i] = 1
        english(bot,message,type,i)

    elif message.text.lower() == 'russian/русский' or message.text.lower() == 'russian':
        Language[i] = 0
        russian(bot,message,type,i)

    elif Language[i] == 1:
        english(bot,message,type,i)

    elif Language[i] == 0:
        russian(bot,message,type,i)

    print('ID=', id, '; Name=', name, '; TYPE=', type, '; i=', i, '; FIND=', find, '; Language=', Language)
    write_user_info('info.xlsx', 'Users')


read_user_info('info.xlsx', 'Users')
bot.polling()